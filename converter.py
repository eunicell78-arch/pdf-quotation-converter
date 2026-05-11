#!/usr/bin/env python3
"""
PDF Quotation to CSV Converter (Universal Version)
다양한 벤더의 견적서 양식을 표준 CSV로 변환

지원 기능:
- 컬럼명 동의어 사전 (Product/Part No./P/N, L/T/Lead Time 등)
- Product 칸 자동 파싱 (제품명, Rated Current, Cable Length, Description)
- SINBON P/N 자동 인식 (제품명이 없을 때)
- Contract/Catalogue 가격 등급 자동 분리
- 병합 셀 자동 채우기
- NRE List 별도 처리
- 머리글 정보 자동 추출 (To/From/Date)
"""

import pdfplumber
import pandas as pd
import re
import sys
import os
from typing import Dict, List, Tuple, Optional
from datetime import datetime

# openpyxl은 엑셀 파일 처리 시에만 import (PDF만 처리하는 환경에서도 동작하도록)
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ============================================================
# 1. 컬럼명 동의어 사전 (벤더마다 다른 컬럼명 통일)
# ============================================================
COLUMN_SYNONYMS = {
    'item': ['item', 'no.', 'no', '#'],
    'product': ['product', 'part no.', 'part no', 'part number', 'part#', 'p/n', 'pn',
                'item name', 'model', 'item description'],
    'description': ['description', 'desc', 'specification', 'spec', 'specs'],
    'delivery_term': ['delivery term', 'delivery', 'incoterm', 'incoterms',
                      'shipping term', 'shipping', 'terms'],
    'moq': ['moq', 'moq (pcs)', 'qty', 'quantity', 'min order', 'minimum order',
            'order qty', "q'ty", 'moq(pcs)'],
    'price': ['unit price', 'unit price (usd)', 'price', 'cost', 'unit cost',
              'unit price(usd)', 'price (usd)'],
    'lt': ['l/t', 'l/t (wks)', 'lead time', 'lt', 'delivery time',
           'l.t.', 'lead-time', 'lt (wks)', 'l/t(wks)'],
    'remark': ['remark', 'remarks', 'note', 'notes', 'comment', 'comments'],
    'cavity': ['cavity', 'cavities'],
    'amount': ['amount', 'total', 'total price', 'subtotal'],
}


# ============================================================
# 2. 패턴 정규식 (모두 반각+전각 콜론 지원)
# ============================================================
RATED_INLINE_PATTERN = re.compile(r'(?i)\brated\s+current\s*[:\uff1a]\s*(.+?)$')
CABLE_INLINE_PATTERN = re.compile(r'(?i)\bcable\s+length\s*[:\uff1a]\s*(.+?)$')
SINBON_PN_PATTERN = re.compile(r'(?i)\bsinbon\s+p/?n\s*[:\uff1a]\s*(.+?)$')
BULLET_PATTERN = re.compile(r'^[-•·–—*]\s*')
KEY_VALUE_PATTERN = re.compile(r'^[A-Za-z][A-Za-z0-9 /&()#-]{1,40}\s*[:\uff1a]')
PRICE_PATTERN = re.compile(r'\$[\d,]+(?:\.\d+)?|[\d,]+\.\d{2,}')


# ============================================================
# 3. 헬퍼 함수
# ============================================================
def safe_get(row, idx, default=''):
    """안전한 리스트 인덱싱"""
    if row is None or idx is None:
        return default
    if not (0 <= idx < len(row)):
        return default
    val = row[idx]
    return val if val is not None else default


def match_column(header_text):
    """헤더 셀 텍스트를 표준 컬럼명으로 매핑"""
    if not header_text:
        return None
    norm = re.sub(r'\s+', ' ', str(header_text).lower().strip())
    norm = norm.rstrip(':').strip()

    for std_name, synonyms in COLUMN_SYNONYMS.items():
        for syn in synonyms:
            if norm == syn:
                return std_name
            # 'unit price (USD)' 같은 변형 처리
            if len(syn) >= 4 and syn in norm:
                rest = norm.replace(syn, '').strip()
                if rest in ['', '(usd)', '(pcs)', '(wks)', '(s)']:
                    return std_name
    return None


def normalize_lt_value(lt_value):
    """L/T 값에 'wks' 단위 일관되게 붙이기"""
    if lt_value is None:
        return ''
    lt_text = str(lt_value).strip()
    if not lt_text:
        return ''
    base = re.sub(r'(?i)\s*(?:wk|wks|week|weeks|w)\.?\s*$', '', lt_text).strip()
    if not base:
        return 'wks'
    return f'{base}wks'


def format_date_to_iso(date_str):
    """다양한 날짜 형식을 yyyy-mm-dd로 변환"""
    if not date_str:
        return ''
    date_str = str(date_str).strip()

    formats = [
        '%b. %d, %Y', '%B. %d, %Y',
        '%b %d, %Y', '%B %d, %Y',
        '%b-%d-%Y', '%B-%d-%Y',
        '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d',
        '%Y-%m-%d', '%Y.%m.%d',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return date_str


# ============================================================
# 4. Product 칸 파싱
# ============================================================
def parse_product_field(product_text):
    """
    Product 칸을 다음 규칙으로 분리:
    1. Rated Current, Cable Length는 별도 추출
    2. 첫 줄이 일반 텍스트면 → Product
    3. 첫 줄이 키:값 형식이고 SINBON P/N이 있으면 → SINBON P/N 값이 Product
    4. 그 외는 첫 줄 그대로
    """
    if not product_text or not str(product_text).strip():
        return '', '', '', ''

    lines = []
    for raw in str(product_text).strip().split('\n'):
        s = raw.strip()
        if s:
            lines.append(BULLET_PATTERN.sub('', s).strip())

    if not lines:
        return '', '', '', ''

    rated_current = ''
    cable_length = ''
    remaining = []

    # Rated/Cable 추출
    for line in lines:
        r = RATED_INLINE_PATTERN.search(line)
        c = CABLE_INLINE_PATTERN.search(line)
        if r and not rated_current:
            rated_current = r.group(1).strip()
            continue
        if c and not cable_length:
            cable_length = c.group(1).strip()
            continue
        remaining.append(line)

    # Product 결정
    product_name = ''
    description_parts = []

    if remaining:
        first = remaining[0]
        if KEY_VALUE_PATTERN.match(first):
            # 키:값 형식 → SINBON P/N 찾기
            sinbon_val = None
            for line in remaining:
                m = SINBON_PN_PATTERN.search(line)
                if m:
                    sinbon_val = m.group(1).strip()
                    break
            if sinbon_val:
                product_name = sinbon_val
                for line in remaining:
                    if not SINBON_PN_PATTERN.search(line):
                        description_parts.append(line)
            else:
                product_name = first
                description_parts.extend(remaining[1:])
        else:
            product_name = first
            description_parts.extend(remaining[1:])

    description = '\n'.join(description_parts).strip()
    return product_name, rated_current, cable_length, description


# ============================================================
# 5. 표 분류 (헤더/메인/NRE)
# ============================================================
def classify_table(table):
    """표가 어떤 종류인지 판단: 'header', 'main', 'nre', 'unknown'"""
    if not table or not table[0]:
        return 'unknown'

    first_row = table[0]
    first_row_text = ' '.join(str(c) for c in first_row if c).lower()

    # 헤더 정보 표 (To/From)
    if 'to:' in first_row_text and 'from:' in first_row_text:
        return 'header'

    # NRE List 표 (Cavity 또는 Amount 컬럼 존재)
    if 'cavity' in first_row_text or 'amount' in first_row_text:
        # 단, Description과 Qty가 있어야 함
        if 'description' in first_row_text or 'qty' in first_row_text:
            return 'nre'

    # 메인 견적표 - 헤더가 있는 경우
    matched = [match_column(c) for c in first_row if c]
    matched = [m for m in matched if m]
    if 'price' in matched and ('moq' in matched or 'product' in matched):
        return 'main'

    # 메인 견적표 - 헤더 없이 데이터가 바로 시작 (견적서 2 page 2 같은 경우)
    if len(first_row) >= 6:
        has_price = any('$' in str(c) for c in first_row if c)
        has_number = any(
            str(c).replace(',', '').replace('.', '').strip().isdigit()
            for c in first_row if c
        )
        if has_price and has_number:
            return 'main_headerless'

    return 'unknown'


# ============================================================
# 6. 컬럼 위치 파악
# ============================================================
def find_column_indices(header_row):
    """헤더 행을 분석해 각 표준 컬럼의 위치 찾기"""
    indices = {}
    for i, cell in enumerate(header_row):
        std = match_column(cell)
        if std and std not in indices:
            indices[std] = i
    return indices


# ============================================================
# 7. 메인 변환기
# ============================================================
class QuotationConverter:
    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.header_info = {}
        self.main_rows = []     # 메인 견적 데이터
        self.nre_rows = []      # NRE List 데이터

    def extract_header_from_table(self, table):
        """To/From/Date 헤더 표에서 정보 추출"""
        for row in table:
            row_text = ' | '.join(str(c) if c else '' for c in row)
            # 모든 셀을 키-값 쌍으로 처리
            for i in range(0, len(row) - 1):
                if not row[i]:
                    continue
                label = str(row[i]).strip().rstrip(':').lower()
                value = str(row[i+1]).strip() if row[i+1] else ''
                if not value:
                    continue
                if label == 'to' and 'customer' not in self.header_info:
                    self.header_info['customer'] = value
                elif label == 'from' and 'planner' not in self.header_info:
                    self.header_info['planner'] = value
                elif label == 'date' and 'date' not in self.header_info:
                    self.header_info['date'] = format_date_to_iso(value)
                elif label == 'attn' and 'attn' not in self.header_info:
                    self.header_info['attn'] = value
                elif label == 'ref' and 'ref' not in self.header_info:
                    self.header_info['ref'] = value

    def extract_header_from_text(self, text):
        """페이지 텍스트에서 헤더 정보 보충"""
        if not text:
            return
        # ADDCOM 같이 표 구조 아닌 양식 대응
        date_match = re.search(r'\b(\d{4}[/-]\d{1,2}[/-]\d{1,2})\b', text)
        if date_match and 'date' not in self.header_info:
            self.header_info['date'] = format_date_to_iso(date_match.group(1))

    def process_main_table(self, table, has_header=True):
        """메인 견적표 처리"""
        if has_header:
            col_idx = find_column_indices(table[0])
            data_rows = table[1:]
        else:
            # 헤더 없는 표 (이전 페이지 표의 연속)
            # 표준 컬럼 순서: Item, Product, Delivery Term, MOQ, Price, L/T, Remark
            col_count = len(table[0]) if table else 0
            if col_count == 7:
                col_idx = {'item': 0, 'product': 1, 'delivery_term': 2, 'moq': 3,
                          'price': 4, 'lt': 5, 'remark': 6}
            elif col_count == 6:
                # 견적서 2 page 2: Item 컬럼 빠짐
                col_idx = {'product': 0, 'moq': 1, 'price': 2, 'lt': 3, 'remark': 4}
                # 실제로 보니 6컬럼인 경우엔 Item이 product에 들어감
                # 데이터 패턴을 보면 [item, product, moq, price, lt, remark]일 수도 있음
                col_idx = self._guess_columns_from_data(table)
            else:
                return
            data_rows = table

        # 빈 행은 위 값 사용 (병합 셀 처리) + 가격 등급 자동 분리
        current = {'item': '', 'product': '', 'delivery_term': '',
                   'moq': '', 'lt': '', 'remark': ''}

        for row in data_rows:
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue

            # 각 컬럼 값 읽기
            new_values = {}
            for key in current.keys():
                if key in col_idx:
                    val = safe_get(row, col_idx[key])
                    val_str = str(val).strip() if val else ''
                    new_values[key] = val_str

            # 값이 있으면 current 업데이트 (병합 셀 보완)
            for key, val in new_values.items():
                if val:
                    current[key] = val

            # Price는 별도로 처리 - 가격이 있는 행만 유효
            price_val = safe_get(row, col_idx.get('price', -1))
            price_str = str(price_val).strip() if price_val else ''

            if not price_str or not PRICE_PATTERN.search(price_str):
                continue

            # 행 추가
            self.main_rows.append({
                'item': current['item'],
                'product': current['product'],
                'delivery_term': current['delivery_term'],
                'moq': current['moq'],
                'price': price_str,
                'lt': current['lt'],
                'remark': current['remark'],
            })

    def _guess_columns_from_data(self, table):
        """헤더 없는 표에서 데이터 패턴으로 컬럼 추정"""
        # 6컬럼: 견적서 2 page 2의 경우 [item, product, moq, price, lt, remark]
        # 또는 [product, moq, price, lt, remark, ?] 가능
        if not table or len(table[0]) < 6:
            return {'item': 0, 'product': 1, 'delivery_term': 2, 'moq': 3,
                    'price': 4, 'lt': 5, 'remark': 6}

        # 가격($)이 들어 있는 컬럼 찾기
        price_col = None
        for c_idx in range(len(table[0])):
            for r in table[:5]:
                if r and c_idx < len(r) and r[c_idx] and '$' in str(r[c_idx]):
                    price_col = c_idx
                    break
            if price_col is not None:
                break

        if price_col is None:
            return {'item': 0, 'product': 1, 'delivery_term': 2, 'moq': 3,
                    'price': 4, 'lt': 5, 'remark': 6}

        # 가격 기준으로 역추정: price 앞에 moq, 그 앞에 delivery_term 또는 product
        result = {'price': price_col}
        if price_col >= 1:
            result['moq'] = price_col - 1
        if price_col >= 2:
            # 견적서 2처럼 delivery_term 없는 경우 → product
            result['product'] = price_col - 2 if price_col >= 2 else 0
        if price_col >= 3:
            result['item'] = 0
            result['product'] = price_col - 2
        if price_col + 1 < len(table[0]):
            result['lt'] = price_col + 1
        if price_col + 2 < len(table[0]):
            result['remark'] = price_col + 2

        return result

    def process_nre_table(self, table):
        """NRE List 표 처리"""
        col_idx = find_column_indices(table[0])

        for row in table[1:]:
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue

            desc = str(safe_get(row, col_idx.get('description', 0))).strip()
            if not desc:
                continue

            cavity = str(safe_get(row, col_idx.get('cavity', -1))).strip()
            qty = str(safe_get(row, col_idx.get('moq', -1))).strip()
            unit_price = str(safe_get(row, col_idx.get('price', -1))).strip()
            lt = str(safe_get(row, col_idx.get('lt', -1))).strip()
            remark = str(safe_get(row, col_idx.get('remark', -1))).strip()

            full_desc = desc
            if cavity:
                full_desc = f"{desc}\nCavity: {cavity}"

            self.nre_rows.append({
                'product': desc,
                'description': full_desc,
                'qty': qty,
                'price': unit_price,
                'lt': lt,
                'remark': remark,
            })

    def convert(self):
        """파일 확장자에 따라 PDF 또는 엑셀로 분기 처리"""
        ext = os.path.splitext(self.pdf_path)[1].lower()
        if ext in ('.xlsx', '.xls', '.xlsm'):
            return self._convert_excel()
        else:
            return self._convert_pdf()

    def _convert_pdf(self):
        """PDF → CSV 변환 메인 로직"""
        with pdfplumber.open(self.pdf_path) as pdf:
            previous_main_columns = None

            for page in pdf.pages:
                tables = page.extract_tables()
                page_text = page.extract_text()

                # 페이지 텍스트로 헤더 보충
                self.extract_header_from_text(page_text)

                for table in tables:
                    if not table:
                        continue

                    kind = classify_table(table)

                    if kind == 'header':
                        self.extract_header_from_table(table)
                    elif kind == 'main':
                        self.process_main_table(table, has_header=True)
                        previous_main_columns = find_column_indices(table[0])
                    elif kind == 'main_headerless':
                        self.process_main_table(table, has_header=False)
                    elif kind == 'nre':
                        self.process_nre_table(table)

        # 가격 등급 자동 분리는 process_main_table에서 이미 처리됨 (Remark 병합 셀로)
        return self._build_dataframe()

    def _convert_excel(self):
        """엑셀 파일을 표 형태로 읽어와 처리

        SINBON 등 엑셀 견적서는 PDF와 컬럼 구조가 거의 동일.
        병합 셀을 풀어서 표준 표 형식(rows of cells)으로 변환한 뒤
        PDF와 동일한 process_main_table / extract_header_from_table을 활용한다.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("엑셀 파일을 처리하려면 openpyxl이 필요합니다. 'pip install openpyxl'로 설치하세요.")

        wb = openpyxl.load_workbook(self.pdf_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._process_excel_sheet(ws)

        return self._build_dataframe()

    def _process_excel_sheet(self, ws):
        """엑셀 시트 하나를 처리해 main_rows와 nre_rows, header_info를 채움"""
        # 1) 시트의 모든 셀 값을 2차원 배열로 가져오기 (병합 셀은 좌상단에만 값 있음)
        max_row = ws.max_row
        max_col = ws.max_column

        # 병합된 셀의 값을 모든 병합 셀에 복사 (PDF 표 추출과 동일한 형태로 만들기)
        merged_value_map = {}
        for merged_range in ws.merged_cells.ranges:
            top_left_val = ws.cell(merged_range.min_row, merged_range.min_col).value
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    if (r, c) != (merged_range.min_row, merged_range.min_col):
                        merged_value_map[(r, c)] = top_left_val

        # 2) 2차원 그리드 만들기
        grid = []
        for r in range(1, max_row + 1):
            row = []
            for c in range(1, max_col + 1):
                if (r, c) in merged_value_map:
                    val = merged_value_map[(r, c)]
                else:
                    val = ws.cell(r, c).value
                # datetime은 문자열로 변환
                if isinstance(val, datetime):
                    val = val.strftime('%Y-%m-%d')
                row.append(val)
            grid.append(row)

        # 3) 헤더 정보 추출 (To/From/Date/Ref/Attn/CC 같은 패턴 찾기)
        self._extract_header_from_grid(grid)

        # 4) 메인 견적표 헤더 행 찾기 (Item, Product, MOQ 등이 같은 행에 나오는 행)
        header_row_idx = self._find_main_table_header_row(grid)
        if header_row_idx is None:
            return

        # 5) 헤더 행과 데이터 행 추출
        header_row = grid[header_row_idx]
        data_rows = []
        for r_idx in range(header_row_idx + 1, len(grid)):
            row = grid[r_idx]
            # 시트 끝부분의 Notes/Currency/서명 영역은 제외
            row_text = ' '.join(str(c) for c in row if c is not None).lower()
            if any(stop in row_text for stop in ['notes:', 'currency:', 'payment terms:',
                                                  'sales supervisor', 'sales engineer', 'form no.']):
                break
            # 행에 데이터가 하나라도 있으면 포함
            if any(c is not None and str(c).strip() != '' for c in row):
                data_rows.append(row)

        # 6) PDF 처리와 동일한 형태로 가공 (None은 빈 문자열로)
        # 또한 가격 컬럼의 숫자값을 $X.XX 형식으로 변환 (PDF처럼 인식되게)
        price_col_idx = None
        for c_idx, h_cell in enumerate(header_row):
            if match_column(h_cell) == 'price':
                price_col_idx = c_idx
                break

        def _format_cell_for_pdf_compat(val, col_idx):
            """엑셀 셀 값을 PDF 표 추출과 동일한 문자열 형태로 변환"""
            if val is None:
                return ''
            # 가격 컬럼의 숫자값은 $ 기호 붙이기 (PDF와 동일하게 인식되도록)
            if col_idx == price_col_idx and isinstance(val, (int, float)):
                return f'${val}'
            return str(val)

        cleaned_header = [str(c) if c is not None else '' for c in header_row]
        cleaned_data = [
            [_format_cell_for_pdf_compat(c, i) for i, c in enumerate(row)]
            for row in data_rows
        ]
        table = [cleaned_header] + cleaned_data

        # 7) 표 종류 판단 후 PDF용 메서드 재사용
        kind = classify_table(table)
        if kind == 'main':
            self.process_main_table(table, has_header=True)
        elif kind == 'main_headerless':
            self.process_main_table(table, has_header=False)
        elif kind == 'nre':
            self.process_nre_table(table)

    def _extract_header_from_grid(self, grid):
        """엑셀 그리드에서 To/From/Date/Attn/CC/Ref 정보 추출"""
        for r_idx, row in enumerate(grid):
            for c_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip().rstrip(':').lower()

                # 라벨 다음 칸에서 값 찾기
                if cell_str in ['to', 'from', 'date', 'attn', 'ref', 'cc']:
                    # 같은 행에서 다음에 나오는 값 있는 셀
                    value = None
                    for next_c in range(c_idx + 1, len(row)):
                        next_val = row[next_c]
                        if next_val is not None and str(next_val).strip():
                            value = str(next_val).strip()
                            break

                    if not value:
                        continue

                    if cell_str == 'to' and 'customer' not in self.header_info:
                        self.header_info['customer'] = value
                    elif cell_str == 'from' and 'planner' not in self.header_info:
                        self.header_info['planner'] = value
                    elif cell_str == 'date' and 'date' not in self.header_info:
                        # 이미 datetime이 yyyy-mm-dd 형식으로 변환되어 있음
                        if re.match(r'^\d{4}-\d{2}-\d{2}', value):
                            self.header_info['date'] = value[:10]
                        else:
                            self.header_info['date'] = format_date_to_iso(value)
                    elif cell_str == 'attn' and 'attn' not in self.header_info:
                        self.header_info['attn'] = value
                    elif cell_str == 'ref' and 'ref' not in self.header_info:
                        self.header_info['ref'] = value

    def _find_main_table_header_row(self, grid):
        """엑셀 그리드에서 메인 견적표의 헤더 행 인덱스 찾기.

        Item/Product/MOQ/Unit Price 같은 표준 컬럼명이 같은 행에 2개 이상 있으면 그 행을 헤더로 봄.
        """
        for r_idx, row in enumerate(grid):
            matched_columns = set()
            for cell in row:
                if cell is None:
                    continue
                std = match_column(cell)
                if std:
                    matched_columns.add(std)
            # 표준 컬럼이 3개 이상 같은 행에 있으면 헤더로 간주
            if len(matched_columns) >= 3 and 'price' in matched_columns:
                return r_idx
        return None

    def _build_dataframe(self):
        """수집된 데이터를 최종 CSV 행으로 변환"""
        rows = []
        date = self.header_info.get('date', '')
        customer = self.header_info.get('customer', '')
        planner = self.header_info.get('planner', '')

        # 메인 견적 행
        for r in self.main_rows:
            product_name, rated, cable, desc = parse_product_field(r['product'])

            moq = r['moq']
            remark = r['remark'].replace('\n', ' ').strip() if r['remark'] else ''
            if 'sample' in moq.lower():
                moq = '1'
                if not remark:
                    remark = 'Sample'

            rows.append({
                'Date': date,
                'Customer': customer,
                'Planner': planner,
                'Product': product_name,
                'Rated Current': rated,
                'Cable Length': cable,
                'Description': desc,
                'Delivery Term': r['delivery_term'].replace('\n', ' ').strip() if r['delivery_term'] else '',
                'MOQ': moq,
                'Price': r['price'],
                'L/T': normalize_lt_value(r['lt']),
                'Remark': remark,
            })

        # NRE List 행
        for r in self.nre_rows:
            rows.append({
                'Date': date,
                'Customer': customer,
                'Planner': planner,
                'Product': r['product'].replace('\n', ' ').strip(),
                'Rated Current': '',
                'Cable Length': '',
                'Description': r['description'],
                'Delivery Term': 'NRE List',
                'MOQ': r['qty'],
                'Price': r['price'],
                'L/T': normalize_lt_value(r['lt']),
                'Remark': r['remark'].replace('\n', ' ').strip() if r['remark'] else '',
            })

        return pd.DataFrame(rows)

    def save_to_csv(self, output_path):
        df = self.convert()
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"✅ 변환 완료: {output_path}")
        print(f"📊 총 행 수: {len(df)}")
        return df


def main():
    if len(sys.argv) < 3:
        print("사용법: python converter.py <입력PDF> <출력CSV>")
        sys.exit(1)

    try:
        converter = QuotationConverter(sys.argv[1])
        converter.save_to_csv(sys.argv[2])
    except Exception as e:
        print(f"❌ 오류: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
